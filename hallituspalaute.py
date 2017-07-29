#!/usr/bin/python
# -*- coding: utf8 -*-

import time
from openpyxl import load_workbook


def kirjaa_hallituspalaute(palaute):
    palautteen_pituus = len(palaute)
    if palautteen_pituus > 250:
        palautettava = "Liian pitkä palaute! Rajoita avautumista ja yritä" \
                       " uudestaan alle 250 merkillä ja " \
                       "/hallituspalaute-komennolla."
        return palautettava

    # Avataan hallitustaulukko.
    wb = load_workbook('hallituspalaute.xlsx')
    worksheet = wb.active
    palautteiden_lkm = worksheet.max_row
    kirjattava = "{} @ {}.{}.{}, {}:{}".format(palaute, time.localtime()[2],
                                               time.localtime()[1],
                                               time.localtime()[0],
                                               time.localtime()[3],
                                               time.localtime()[4])
    worksheet['A{}'.format(palautteiden_lkm + 1)] = kirjattava
    # Palautteen kirjaamisen jalkeen palautteita on yllattaen yksi enemman
    palautteiden_lkm += 1

    # Jos palautteen kirjaamisen jalkeen on ylitetty raja
    #  palautteiden maarassa, siistitaan lokia poistamalla
    # vanhempi lokin puolikas.
    if palautteiden_lkm >= 101:
        print("Palautteiden maksimimaara ylitetty, siistitaan lokia..")
        for rivi in range(52, palautteiden_lkm + 1):
            worksheet['A{}'.format(rivi-50)] = \
                worksheet['A{}'.format(rivi)].value
            worksheet['A{}'.format(rivi)].value = None
    # Lopuksi tallennetaan muutokset tai ilmoitetaan virheesta
    try:
        wb.save('hallituspalaute.xlsx')
        palautettava = "Kiitos palautteestasi!"
    except IOError or PermissionError:
        print("Ei voitu tallentaa palautetta: '" + kirjattava + "'")
        palautettava = "Jokin meni pieleen, ei voitu tallentaa palautetta."

    return palautettava


def raportti():
    # Avataan hallitustaulukko.
    wb = load_workbook('hallituspalaute.xlsx')
    worksheet = wb.active
    palautteiden_lkm = worksheet.max_row
    # Maaritetaan alaraja tulostettaville raporteille
    if palautteiden_lkm <= 11:
        alaraja = 2
    else:
        alaraja = palautteiden_lkm - 9
    # Palautetaan 10 (tai alle) viimeisinta palautetta
    palautettava = "10 viimeisintä palautetta:\n\n"
    try:
        for i in range(alaraja, palautteiden_lkm + 1):
            palautettava = palautettava + worksheet['A{}'.format(i)]\
                .value.encode('utf8') + "\n\n"
    except TypeError:
        print("Koodekkivirhe, olet windowsilla")
    return palautettava
