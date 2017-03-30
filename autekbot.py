import json
import time
import random

import requests
from requests import exceptions
from openpyxl import load_workbook
import komentotarkistin


###############################################################################
#  --------------------------------  @Autekbot  ----------------------------  #
#  --------------------------------  Samu Ampio ----------------------------  #
###############################################################################

TOKEN = "308527009:AAFPg5p53k-I0iYuWJNU-eDJTRGutg2Xx_8"
URL = "https://api.telegram.org/bot{}/".format(TOKEN)
WHITELIST = {21942357, 152093174, 39307350, 141787534}


def get_url(url):
    onnistui = False
    # Niin kauan kun
    while not onnistui:
        try:
            response = requests.get(url)
            content = response.content.decode("utf8")
        except requests.exceptions.ConnectionError:
            print("sleeping..")
            time.sleep(10)
            onnistui = False
            continue
        return content


def get_json_from_url(url):
    content = get_url(url)
    js = json.loads(content)
    return js


def get_updates(offset=None):
    url = URL + "getUpdates"
    if offset:
        url += "?offset={}".format(offset)
    js = get_json_from_url(url)
    return js


def get_last_chat_id_and_text(updates):
    num_updates = len(updates["result"])
    last_update = num_updates - 1
    chat_id = updates["result"][last_update]["message"]["chat"]["id"]
    try:
        text = updates["result"][last_update]["message"]["text"]
        return text, chat_id
    except KeyError or UnicodeEncodeError:
        return "I'm blind, you need to speak. U_U", chat_id


def send_message(text, chat_id):
    try:
        url = URL + "sendMessage?text={}&chat_id={}".format(text, chat_id)
    except UnicodeEncodeError:
        url = URL + "sendMessage?text=I don't know " \
                    "how to say that.. :o&chat_id={}".format(chat_id)
    get_url(url)
    return


def get_last_update_id(updates):
    update_ids = []
    for update in updates["result"]:
        update_ids.append(int(update["update_id"]))
    max_id = str(max(update_ids))
    return int(max_id)


# echo_all(paivitykset) kay lapi kaikki serverin puskurissa olevat viestit
# ja kasittelee ne tavalla tai toisella.
def echo_all(paivitykset):
    # Kaydaan paivityslista lapi.
    for update in paivitykset["result"]:
        # Tallennetaan viestit ja chat_id:t muuttujiin.
        try:
            teksti = update["message"]["text"]
        except KeyError or UnicodeEncodeError:
            continue
        chat = update["message"]["chat"]["id"]
        vastausteksti, chat = komentotarkistin.tarkista_komento(update, teksti, chat)
        send_message(vastausteksti, chat)
    return


# hallitusnakki() lukee botin tyokansiossa olevan hallitus-excelin
# ja arpoo nimien joukosta yhden seka palauttaa sen.
def hallitusnakki():
    # Avataan hallitustaulukko.
    worksheet = load_workbook('hallitus.xlsx').active
    # Tarkistetaan montako nimea on listassa
    # ja arvotaan luku 1 ja maksimin valilta.
    arpa = random.randrange(1, worksheet.max_row)
    # Kaivetaan arpalukua vastaava yhteystieto ja palautetaan se.
    nakkinimi = worksheet['A{}'.format(arpa)].value
    return nakkinimi


def main():
    edellinen_paiv = None
    while True:
        # Haetaan paivitykset serverilta.
        paivitykset = get_updates(edellinen_paiv)
        if len(paivitykset["result"]) > 0:
            edellinen_paiv = get_last_update_id(paivitykset) + 1
            echo_all(paivitykset)
        # Odotetaan hyvan maun nimissa vahan aikaa
        time.sleep(0.5)

main()
