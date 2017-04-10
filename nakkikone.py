import random
from openpyxl import load_workbook


# hallitusnakki() lukee botin tyokansiossa olevan hallitus-excelin
# ja arpoo nimien joukosta yhden seka palauttaa sen.
def hallitusnakki():
    # Avataan hallitustaulukko.
    worksheet = load_workbook('hallitus.xlsx').active
    # Tarkistetaan montako nimea on listassa
    # ja arvotaan luku 1 ja maksimin valilta.
    arpa = random.randrange(1, worksheet.max_row)
    # Kaivetaan arpalukua vastaava yhteystieto ja palautetaan se.
    nakkinimi = worksheet['A{}'.format(arpa)].value

    return nakkinimi
