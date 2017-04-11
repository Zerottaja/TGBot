#!/usr/bin/python
# -*- coding: utf8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook


# lisaa_kayttaja() tutkii onko lisaysta pyytaneen kayttajan id jo listalla,
# ja mikali ei ole, lisaa sen listaan seka excel-tiedostoon,
# etta valimuistin muuttujalistaan
def lisaa_kayttaja(kayttaja_id):
    # Jos id ei ole viela listassa
    if not etsi_kayttaja(kayttaja_id):
        try:
            # Avataan id-taulukko.
            wb = load_workbook('WHITELIST.xlsx')
        except FileNotFoundError:
            wb = Workbook()
        worksheet = wb.active
        kayttajien_lkm = worksheet.max_row
        # lisataan id listan loppuun
        worksheet['A{}'.format(kayttajien_lkm + 1)].value = kayttaja_id
        wb.save('WHITELIST.xlsx')
        # paivitetaan myos valimuisti
        paivita_muuttujalista()

    return


# etsi_kayttaja() etsii id:n perusteella
# excel-tiedostosta osumaa ja palauttaa loytyiko sita
def etsi_kayttaja(kayttaja_id):
    try:
        # Avataan id-taulukko.
        wb = load_workbook('WHITELIST.xlsx')
    except FileNotFoundError:
        return False
    worksheet = wb.active
    korkein_rivi = worksheet.max_row
    # jos korkein rivinumero on 1, listassa ei ole alkioita
    if korkein_rivi == 1:
        return False

    # aloitetaan pyyhkaiseva haku 2. rivilta
    i = 2
    while i <= korkein_rivi:
        # jos rivin arvo matchaa haetun id:n kanssa, palautetaan True
        if kayttaja_id == worksheet['A{}'.format(i)].value:
            return True
        # jos rivin arvo ei matchaa mutta on tyhja, siirretaan
        # listan viimeinen alkio tahan koloon listan siistimiseksi
        elif worksheet['A{}'.format(i)].value is None:
            worksheet['A{}'.format(i)].value \
                = worksheet['A{}'.format(korkein_rivi)].value
            # kun viimeinen on kopioitu koloon, se voidaan poistaa lopusta
            worksheet['A{}'.format(korkein_rivi)].value = None
            # paivitetaan rivimaara
            korkein_rivi = worksheet.max_row
            wb.save('WHITELIST.xlsx')
        i += 1

    return False


# paivita_muuttujalista() resetoi valimuistissa olevan whitelistin
# ja rakentaa sen uudelleen excel-tiedoston perusteella
def paivita_muuttujalista():
    try:
        # Avataan id-taulukko.
        wb = load_workbook('WHITELIST.xlsx')
    except FileNotFoundError:
        return
    worksheet = wb.active
    korkein_rivi = worksheet.max_row
    # tyhjenetaan globaali muuttuja WHITELIST
    import autekbot
    del autekbot.WHITELIST[:]
    # rakennetaan lista uudestaan
    for rivi in range(2, korkein_rivi + 1):
        autekbot.WHITELIST.append(worksheet['A{}'.format(rivi)].value)
    return
